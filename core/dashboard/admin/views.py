from django.views.generic import (
    TemplateView,
    ListView,
    DetailView,
    UpdateView,
    CreateView,
    DeleteView
)
from django.db.models import F, DecimalField, ExpressionWrapper
from django.db.models.functions import Round
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.views import PasswordChangeView
from django.contrib import messages
from django.db.models import Sum
from django.contrib.auth.models import User
from django.shortcuts import redirect
from django.db.models import ProtectedError
from django.http import HttpResponse
from django.views import View
import openpyxl
from openpyxl.styles import Font


from website.models import NewsLetterModel
from ..permissions import HasAdminAccessPermission
from .forms import (
    AdminPasswordChangeForm,
    AdminProductDetailEditFrom,
)
from ..jalali_utils import (
    PERSIAN_MONTH_NAMES_FA,
    get_current_jalali_ymd,
    jalali_month_start_to_aware_gregorian,
    jalali_month_range_to_aware_gregorian,
)
from shop.models import (
    ProductModel,
    ProductCategoryModel,
    ProductStatusType
)
from order.models import (
    OrderModel,
    OrderStatusType
)


class AdminDashboardHomeView(HasAdminAccessPermission,TemplateView):
    template_name = 'dashboard/admin/main/main.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        jy, jm, jd = get_current_jalali_ymd()
        month_start = jalali_month_start_to_aware_gregorian(jy, jm)

        successful_orders = OrderModel.objects.filter(status=OrderStatusType.succes.value)

        
        context["total_sales"] = self.get_sales_sum(successful_orders)      
        context["month_sales"] = self.get_sales_sum(successful_orders, since=month_start)

        context["total_orders"] = OrderModel.objects.count()
        context["month_orders"] = OrderModel.objects.filter(created_date__gte=month_start)[:5]
        
        context['total_customers'] = User.objects.count()

        context['total_products'] = ProductModel.objects.count()
        
        # context["today_jalali_full"] = today_jalali.strftime("%Y/%m/%d")
        context["current_month_name"] = PERSIAN_MONTH_NAMES_FA[jm - 1]

        # cancelled=>faild pending complete delivered=>succes
        
        
        return context

    def get_sales_sum(self, queryset, since=None):
        if since:
            queryset = queryset.filter(created_date__gte=since)
        result = queryset.aggregate(total=Sum("total_price"))
        return result["total"] or 0
    
    
class AdminSecurityEditView(LoginRequiredMixin, HasAdminAccessPermission, SuccessMessageMixin, PasswordChangeView):
    template_name = 'dashboard/admin/settings/settings.html'
    form_class = AdminPasswordChangeForm
    success_url = reverse_lazy("dashboard:admin:security-edit")
    success_message = 'رمز با موفقیت عوض شد'
    
    
class AdminOrdersView(LoginRequiredMixin, HasAdminAccessPermission, SuccessMessageMixin, ListView):
    template_name = 'dashboard/admin/orders/orders.html'
    context_object_name = 'orders'
    paginate_by = 5
    
    def get_queryset(self):
        queryset = OrderModel.objects.all()
        
        if q:= self.request.GET.get('q'):
            queryset = queryset.filter(user__user_profile__last_name__icontains=q)
        
        order_status = self.request.GET.get('order-status')
        if order_status == 'complete':
            queryset = queryset.filter(status=OrderStatusType.complete.value)
        if order_status == 'pending':
            queryset = queryset.filter(status=OrderStatusType.pending.value)
        if order_status == 'succes':
            queryset = queryset.filter(status=OrderStatusType.succes.value)
        if order_status == 'faild':
            queryset = queryset.filter(status=OrderStatusType.faild.value)
        if order_status == 'all':
            pass
        
        
        jalali_year = self.request.GET.get('jalali_year')
        jalali_month = self.request.GET.get('jalali_month')

        if jalali_year and jalali_month:
            try:
                jy = int(jalali_year)
                jm = int(jalali_month)
                start, end = jalali_month_range_to_aware_gregorian(jy, jm)
                queryset = queryset.filter(created_date__gte=start, created_date__lt=end)
            except (ValueError, TypeError):
                pass
        elif jalali_year:
            try:
                jy = int(jalali_year)
                start = jalali_month_start_to_aware_gregorian(jy, 1)
                end = jalali_month_start_to_aware_gregorian(jy + 1, 1)
                queryset = queryset.filter(created_date__gte=start, created_date__lt=end)
            except (ValueError, TypeError):
                pass

        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        jy, jm, jd = get_current_jalali_ymd()
        month_start = jalali_month_start_to_aware_gregorian(jy, jm)
        
        context["total_orders"] = OrderModel.objects.count()
        context["month_orders"] = OrderModel.objects.filter(created_date__gte=month_start)[:5]
        context['success_orders'] = OrderModel.objects.filter(status=OrderStatusType.succes.value).count()
        
        context['pending_orders'] = OrderModel.objects.filter(status=OrderStatusType.pending.value).count()
        context['complete_orders'] = OrderModel.objects.filter(status=OrderStatusType.complete.value).count()
        context['faild_orders'] = OrderModel.objects.filter(status=OrderStatusType.faild.value).count()

        order_status = self.request.GET.get('order-status')
        context['order_filter'] = order_status
        
        context['jalali_months'] = list(enumerate(PERSIAN_MONTH_NAMES_FA, start=1))
        context['jalali_years'] = range(jy - 5, jy + 1)   # ۵ سال اخیر تا سال جاری

        context['selected_jalali_year'] = self.request.GET.get('jalali_year', '')
        context['selected_jalali_month'] = self.request.GET.get('jalali_month', '')

        return context

    def get_sales_sum(self, queryset, since=None):
        if since:
            queryset = queryset.filter(created_date__gte=since)
        result = queryset.aggregate(total=Sum("total_price"))
        return result["total"] or 0


class AdminOrdersDetailView(LoginRequiredMixin, HasAdminAccessPermission, SuccessMessageMixin, DetailView):
    template_name = 'dashboard/admin/orders/orders-detail.html'
    queryset = OrderModel.objects.all()
    context_object_name = 'order'
    
    

class AdminProductsListView(LoginRequiredMixin, HasAdminAccessPermission, ListView):
    template_name = 'dashboard/admin/products/products.html'
    context_object_name = "products"
    paginate_by = 5

    def get_queryset(self):
        queryset = ProductModel.objects.all().annotate(
            final_price=Round(ExpressionWrapper(
                 F("price") - (F("price") * F("discount_percent") / 100),
                 output_field=DecimalField()
            ))
        )
        
        if q := self.request.GET.get('q'):
            queryset = queryset.filter(title__icontains=q)        
        try:
            if min_price := self.request.GET.get('min_price'):
                min_price = int(min_price)
                queryset = queryset.filter(final_price__gte=min_price)
        except (ValueError, TypeError):
            pass 
        try:
            if max_price := self.request.GET.get('max_price'):
                max_price = int(max_price)
                queryset = queryset.filter(final_price__lte=max_price)
        except (ValueError, TypeError):
            pass
        
        filter_by = self.request.GET.get('filter-by')
        
        if filter_by == 'cheep_to_exp':
            queryset = queryset.order_by('final_price')        
        elif filter_by == 'exp_to_cheep':
            queryset = queryset.order_by('-final_price')        
        elif filter_by == 'new':
            queryset = queryset.order_by('-created_date')
        
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_products'] = ProductModel.objects.count()
        context['published_products'] = ProductModel.objects.filter(status=ProductStatusType.publish.value).count()
        context['drafted_products'] = ProductModel.objects.filter(status=ProductStatusType.draft.value).count()

        context['total_categories'] = ProductCategoryModel.objects.count()

        total_price_of_products = sum(product.price for product in ProductModel.objects.all())
        
        context['total_price_of_products'] = total_price_of_products    
        
        return context


class AdminProductDetailEditView(LoginRequiredMixin, HasAdminAccessPermission, SuccessMessageMixin, UpdateView):
    template_name = 'dashboard/admin/products/edit-products.html'
    success_message = 'ویرایش فرم با موفقیت انجام شد'
    queryset = ProductModel.objects.all()
    form_class = AdminProductDetailEditFrom
    context_object_name = 'product'

    def get_success_url(self):
        return reverse_lazy('dashboard:admin:edit-products', kwargs={'pk': self.get_object().pk})


class AdminAddProductView(LoginRequiredMixin, HasAdminAccessPermission, SuccessMessageMixin, CreateView):
    template_name = "dashboard/admin/products/add-product.html"
    success_message = 'افزودن محصول با موفقیت انجام شد'
    queryset = ProductModel.objects.all()
    form_class = AdminProductDetailEditFrom
    context_object_name = 'product'
    
    def form_valid(self, form):
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('dashboard:admin:add-product')


class AdminProductsDeleteView(LoginRequiredMixin, HasAdminAccessPermission, DeleteView):
    success_url = reverse_lazy("dashboard:admin:products-list")
    queryset = ProductModel.objects.all()
    
    def form_valid(self, form):
            try:
                self.object.delete()
                messages.success(self.request, 'محصول با موفقیت حذف شد')
                return redirect(self.success_url)
            except ProtectedError:
                messages.error(
                    self.request, 
                    'این محصول قابل حذف نیست، چون قبلاً در یک یا چند سفارش استفاده شده است.'
                )
                return redirect(self.success_url)
    
    
class AdminCustomersView(TemplateView):
    template_name = 'dashboard/admin/customers/customers.html'
    
    

class NewsletterExportView(LoginRequiredMixin, HasAdminAccessPermission, View):
    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "خبرنامه"

        # هدر ستون‌ها
        sheet['A1'] = 'ردیف'
        sheet['B1'] = 'شماره تلفن'
        sheet['A1'].font = Font(bold=True)
        sheet['B1'].font = Font(bold=True)

        # داده‌ها
        queryset = NewsLetterModel.objects.all().order_by('id')
        for row_num, entry in enumerate(queryset, start=2):
            sheet.cell(row=row_num, column=1, value=row_num - 1)
            sheet.cell(row=row_num, column=2, value=entry.phone_number)

        # عرض ستون‌ها رو کمی بزرگ‌تر می‌کنیم
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="newsletter_subscribers.xlsx"'
        workbook.save(response)
        return response    
    
    
    
    
    